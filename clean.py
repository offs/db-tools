import csv
import json
import sqlite3
import os
from openpyxl import load_workbook

def get_delimiter(file_path):
    with open(file_path, 'r') as file:
        dialect = csv.Sniffer().sniff(file.read(1024))
        delimiter = dialect.delimiter
        if delimiter is None:
            print(f"Could not detect delimiter, defaulting to comma (,)")
            delimiter = ','
        return delimiter

def get_key(row, name_col):
    email = next((value.strip() for value in row if '@' in value), None)
    if email:
        return email
    else:
        values = [value.strip() for i, value in enumerate(row) if i >= name_col and value.strip() and not any(char.isdigit() for char in value)]
        return ' '.join(values[:2])
    
def get_name_column(headers):
    name_keywords = ['name', 'username', 'email', 'full name', 'user']
    first_name_index = last_name_index = -1
    
    for i, header in enumerate(headers):
        header_lower = header.lower()
        if any(keyword in header_lower for keyword in name_keywords):
            return i
        if 'first' in header_lower and 'name' in header_lower:
            first_name_index = i
        if 'last' in header_lower and 'name' in header_lower:
            last_name_index = i
    
    if first_name_index != -1 and last_name_index != -1:
        return (first_name_index, last_name_index)
    
    return 0  # Default to first column if no match found

def read_csv(file_path):
    data = {}
    delimiter = get_delimiter(file_path)
    
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=delimiter)
        headers = next(reader, None)
        
        if headers:
            name_col = get_name_column(headers)
            print(f"Headers: {headers}")
            for row_num, row in enumerate(reader, 2):
                if len(row) > name_col:
                    key = get_key(row, name_col)
                    if key:
                        values = [row[name_col].strip()] + [value.strip() for i, value in enumerate(row) if i != name_col and value.strip() and value != key]
                        if values:
                            if key not in data:
                                data[key] = []
                            data[key] = list(dict.fromkeys(data[key] + values))
                    else:
                        print(f"Skipped: Invalid identifier (Row {row_num})")
                else:
                    print(f"Skipped: Row has insufficient columns (Row {row_num})")
        else:
            file.seek(0)
            for row_num, row in enumerate(reader, 1):
                key = get_key(row, 0)
                if key:
                    values = [value.strip() for value in row if value.strip()]
                    if values:
                        if key not in data:
                            data[key] = []
                        data[key] = list(dict.fromkeys(data[key] + values))
                else:
                    print(f"Skipped: Invalid identifier (Row {row_num})")
    
    print(f"Finished processing. Total entries: {len(data)}")
    return data


def read_sql(file_path):
    data = {}
    conn = sqlite3.connect(file_path)
    cursor = conn.cursor()
    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    table_name = cursor.fetchone()[0]
    
    cursor.execute(f"SELECT * FROM {table_name}")
    headers = [description[0] for description in cursor.description]
    name_col = get_name_column(headers)
    
    print(f"Headers: {headers}")
    
    for row_num, row in enumerate(cursor.fetchall(), 1):
        if any(row):
            if len(row) > name_col:
                key = get_key(row, name_col)
                if key:
                    values = [str(row[name_col]).strip()] + [str(value).strip() for i, value in enumerate(row) if i != name_col and str(value).strip() and str(value) != key]
                    if values:
                        if key not in data:
                            data[key] = []
                        data[key] = list(dict.fromkeys(data[key] + values))
                else:
                    print(f"Skipped: Invalid identifier (Row {row_num})")
            else:
                print(f"Skipped: Row has insufficient columns (Row {row_num})")
    
    conn.close()
    print(f"Finished processing. Total entries: {len(data)}")
    return data

def read_txt(file_path):
    data = {}
    delimiter = get_delimiter(file_path)
    
    with open(file_path, 'r', encoding='utf-8') as txtfile:
        for line_num, line in enumerate(txtfile, 1):
            parts = line.strip().split(delimiter)
            if parts:
                key = get_key(parts, 0)
                if key:
                    values = [parts[0].strip()] + [value.strip() for value in parts[1:] if value.strip() and value != key]
                    if values:
                        if key not in data:
                            data[key] = []
                        data[key] = list(dict.fromkeys(data[key] + values))
                else:
                    print(f"Skipped: Invalid identifier (Line {line_num})")
            else:
                print(f"Skipped: Empty line (Line {line_num})")
                
    print(f"Finished processing. Total entries: {len(data)}")
    return data

def read_xlsx(file_path):
    data = {}
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if headers:
        name_col = get_name_column(headers)
        print(f"Headers: {headers}")
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if len(row) > name_col:
                key = get_key(row, name_col)
                if key:
                    values = [str(value).strip() for i, value in enumerate(row) if i != name_col and str(value).strip()]
                    if values:
                        if key not in data:
                            data[key] = []
                        data[key].extend(values)
                else:
                    print(f"Skipped: Invalid identifier (Row {row_num})")
            else:
                print(f"Skipped: Row has insufficient columns (Row {row_num})")
    else:
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            key = get_key(row, 0)
            if key:
                values = [str(value).strip() for value in row if str(value).strip()]
                if values:
                    if key not in data:
                        data[key] = []
                    data[key].extend(values)
            else:
                print(f"Skipped: Invalid identifier (Row {row_num})")
    
    print(f"Finished processing. Total entries: {len(data)}")
    return data

def write_json(data, output_file):
    with open(output_file, 'w') as jsonfile:
        json.dump(data, jsonfile, indent=2)

def main():
    input_file = input("Enter the input file path: ")
    output_file = os.path.splitext(input_file)[0] + 'out.json'
    
    file_type = os.path.splitext(input_file)[1].lower()
    
    if file_type == '.csv':
        data = read_csv(input_file)
    elif file_type == '.db' or file_type == '.sqlite':
        data = read_sql(input_file)
    elif file_type == '.txt':
        data = read_txt(input_file)
    elif file_type == '.xlsx':
        data = read_xlsx(input_file)
    else:
        print("Unsupported file type")
        return
    
    write_json(data, output_file)
    print(f"Data has been written to {output_file}")

if __name__ == "__main__":
    main()